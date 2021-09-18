# -*- coding: utf-8 -*-
"""export-pq.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1SW9hEX65Y9Y620vp49RMPNvree19PvDJ
"""

# !git clone https://github.com/nuvita97/project-quotation

from google.colab import files
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Font
import pandas as pd

uploaded = files.upload()

# Get the filename
filename = next(iter(uploaded))

df = pd.read_excel(filename, header=None)
# df

# Keep data needed
table = df.iloc[6:-4, [0,1,2,3,4,5,9]]
# table

list = table[1].to_list()

# Read Lookup file in DataFrame
df2 = pd.read_excel('project-quotation/lookup.xlsx').fillna('')
# df2

# Change columns in Lookup file to Lists

initial = df2['initial'].to_list()
# initial = [str(i) for i in initial]
vn_name = df2['vn_name'].to_list()
# vn_name = [str(i) for i in vn_name]
origin = df2['origin'].to_list()
unit = df2['unit'].to_list()

# Loop & Create new list for name
list_name = []

for item in list:
  for i in range(len(initial)):
    if len(vn_name[i]) != 0:
      item = item.replace( initial[i], vn_name[i] )
  list_name.append(item)

# type(list_name)

# Loop & Create new list for origin

list_origin = []
item_origin = ''

for item in list:
  for i in range(len(initial)):
    if initial[i] in item:
      if len(origin[i]) != 0:
        item_origin = origin[i]
  list_origin.append(item_origin)

# type(list_origin)

# Loop & Create new list for unit 

list_unit = []
item_unit = ''

for item in list:
  for i in range(len(initial)):
    if initial[i] in item:
      if len(unit[i]) != 0:
        item_unit = unit[i]
  list_unit.append(item_unit)

# list_unit

# Add new columns for table to match Template

table[1] = list_name
table[2] = list_origin
table[3] = list_unit

table.insert(2, '1', '')
table.insert(2, '2', '')
table.insert(2, '3', '')
table.insert(6, '4', '')

# table

#Use library dataframe_to_rows
rows = dataframe_to_rows(table, index=False, header=False)

wb = load_workbook('project-quotation/template-baogia.xlsx')
ws = wb.active

first_row = 17

# Copy from DataFrame to Excel Template
for r_idx, row in enumerate(rows, 1):
  for c_idx, value in enumerate(row, 1):
    ws.cell(row=r_idx + first_row, column=c_idx, value=value)

# Add new name for created file
new_filename = filename[:-5] + '_modified.xlsx'

wb.save(new_filename)

# rows = dataframe_to_rows(table, index=False, header=False)

# wb2 = Workbook()
# ws2 = wb2.active

# for r_idx, row in enumerate(rows, 1):
#   for c_idx, value in enumerate(row, 1):
#     ws2.cell(row=r_idx, column=c_idx, value=value)


# # for i, line in enumerate(table):
# #   for k, val in enumerate(line):
# #     ws2.cell(row = i + first_row, column = k+1).value = val

# wb2.save('test.xlsx')